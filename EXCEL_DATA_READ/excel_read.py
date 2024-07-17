import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('Excel_data/Reliance Industr.xlsx')

# Get a list of all sheet names
sheet_names = wb.sheetnames

# Iterate over each sheet
for sheet_name in sheet_names:
    sheet = wb[sheet_name]
    print(f"Sheet: {sheet_name}")
    for row in sheet.iter_rows(values_only=True):
        print(row)
